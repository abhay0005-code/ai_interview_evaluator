from pathlib import Path
import csv
import json
from datetime import datetime

REPORT_DIR = Path(__file__).resolve().parent / "data" / "reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)

def _safe(v):
    return "" if v is None else str(v)

def build_report(session, attempts):
    scores = [float(a.get("score") or 0) for a in attempts if a.get("score") is not None]
    avg = round(sum(scores) / len(scores), 2) if scores else 0
    if not attempts:
        overall_feedback = "No answers were submitted, so an overall assessment is not available."
    else:
        level = ("strong interview performance" if avg >= 8 else
                 "solid performance with some areas to strengthen" if avg >= 6 else
                 "a foundation that needs further preparation")
        suggestions = []
        for attempt in attempts:
            suggestion = (attempt.get("suggestions") or attempt.get("feedback") or "").strip()
            if suggestion and suggestion not in suggestions:
                suggestions.append(suggestion)
        overall_feedback = f"The candidate showed {level} (overall score: {avg:.1f}/10)."
        if suggestions:
            overall_feedback += " Focus on: " + " ".join(suggestions[:3])
    return {
        "candidate": session.get("candidate_name", session.get("candidate", "")),
        "candidate_email": session.get("candidate_email", ""),
        "target_role": session.get("target_role", ""),
        "experience_years": session.get("experience_years", ""),
        "section": session.get("category", ""),
        "llm_provider": session.get("provider", ""),
        "llm_model": session.get("model", ""),
        "start_time": session.get("started_at", session.get("start_time", "")),
        "end_time": session.get("ended_at", session.get("end_time", "")),
        "status": session.get("status", ""),
        "overall_score": avg,
        "overall_feedback": overall_feedback,
        "questions_attempted": len(attempts),
        "attempts": attempts,
    }

def export_json(report):
    path = REPORT_DIR / f"interview_{datetime.now():%Y%m%d_%H%M%S}.json"
    path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")
    return str(path)

def export_csv(report):
    path = REPORT_DIR / f"interview_{datetime.now():%Y%m%d_%H%M%S}.csv"
    fields = [
        "question", "answer", "score", "correctness",
        "completeness", "technical_depth", "clarity",
        "response_time", "feedback"
    ]
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for a in report["attempts"]:
            ev = a.get("evaluation") or {}
            writer.writerow({
                "question": a.get("question", ""),
                "answer": a.get("answer", ""),
                "score": ev.get("score", a.get("score", "")),
                "correctness": ev.get("correctness", ""),
                "completeness": ev.get("completeness", ""),
                "technical_depth": ev.get("technical_depth", a.get("depth", "")),
                "clarity": ev.get("clarity", ""),
                "response_time": a.get("response_time_seconds", a.get("response_time", "")),
                "feedback": ev.get("feedback", ev.get("improvement", "")),
            })
    return str(path)

def export_excel(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    path = REPORT_DIR / f"interview_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Candidate", report["candidate"]),
        ("Email", report.get("candidate_email", "")),
        ("Target Role", report.get("target_role", "")),
        ("Experience", report.get("experience_years", "")),
        ("Section", report["section"]),
        ("LLM Provider", report["llm_provider"]),
        ("LLM Model", report["llm_model"]),
        ("Start", report["start_time"]),
        ("End", report["end_time"]),
        ("Status", report["status"]),
        ("Overall Score", report["overall_score"]),
        ("Overall Feedback", report.get("overall_feedback", "")),
        ("Questions Attempted", report["questions_attempted"]),
    ]
    for r, row in enumerate(rows, 1):
        ws.cell(r, 1, row[0]).font = Font(bold=True)
        ws.cell(r, 2, row[1])

    q = wb.create_sheet("Questions & Answers")
    headers = [
        "Question", "Answer", "Score", "Correctness",
        "Completeness", "Technical Depth", "Clarity",
        "Response Time", "Feedback"
    ]
    q.append(headers)
    for c in q[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
        c.alignment = Alignment(wrap_text=True)
    for a in report["attempts"]:
        ev = a.get("evaluation") or {}
        q.append([
            a.get("question", ""),
            a.get("answer", ""),
            ev.get("score", a.get("score", "")),
            ev.get("correctness", ""),
            ev.get("completeness", ""),
            ev.get("technical_depth", a.get("depth", "")),
            ev.get("clarity", ""),
            a.get("response_time_seconds", a.get("response_time", "")),
            ev.get("feedback", ev.get("improvement", "")),
        ])
    for col in q.columns:
        q.column_dimensions[col[0].column_letter].width = min(
            max(15, max(len(str(x.value or "")) for x in col) + 2), 55
        )
    q.freeze_panes = "A2"
    wb.save(path)
    return str(path)

def export_pdf(report):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

    path = REPORT_DIR / f"interview_{datetime.now():%Y%m%d_%H%M%S}.pdf"
    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        rightMargin=14*mm, leftMargin=14*mm,
        topMargin=14*mm, bottomMargin=14*mm
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("AI Interview Evaluation Report", styles["Title"]),
        Spacer(1, 8),
    ]
    summary = [
        ["Candidate", _safe(report["candidate"])],
        ["Email", _safe(report.get("candidate_email", ""))],
        ["Target Role", _safe(report.get("target_role", ""))],
        ["Experience", _safe(report.get("experience_years", ""))],
        ["Section", _safe(report["section"])],
        ["LLM", f'{_safe(report["llm_provider"])} / {_safe(report["llm_model"])}'],
        ["Status", _safe(report["status"])],
        ["Overall Score", f'{report["overall_score"]}/10'],
        ["Overall Feedback", _safe(report.get("overall_feedback", ""))],
        ["Questions Attempted", _safe(report["questions_attempted"])],
    ]
    t = Table(summary, colWidths=[42*mm, 130*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (0,-1), colors.HexColor("#EAF2F8")),
        ("FONTNAME", (0,0), (0,-1), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.4, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
    ]))
    story += [t, Spacer(1, 12)]

    for i, a in enumerate(report["attempts"], 1):
        ev = a.get("evaluation") or {}
        story.append(Paragraph(f"Question {i}: {_safe(a.get('question'))}", styles["Heading3"]))
        story.append(Paragraph(f"<b>Candidate Answer:</b> {_safe(a.get('answer'))}", styles["BodyText"]))
        story.append(Spacer(1, 4))
        metrics = [
            ["Score", ev.get("score", a.get("score", "")),
             "Correctness", ev.get("correctness", ""),
             "Completeness", ev.get("completeness", "")],
            ["Depth", ev.get("technical_depth", a.get("depth", "")),
             "Clarity", ev.get("clarity", ""),
             "Response Time", a.get("response_time_seconds", a.get("response_time", ""))],
        ]
        mt = Table(metrics, colWidths=[24*mm, 24*mm, 30*mm, 24*mm, 30*mm])
        mt.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#F4F6F7")),
        ]))
        story += [mt, Spacer(1, 5)]
        feedback = ev.get("feedback") or ev.get("improvement") or ""
        story.append(Paragraph(f"<b>Feedback:</b> {_safe(feedback)}", styles["BodyText"]))
        story.append(Spacer(1, 12))

    doc.build(story)
    return str(path)
